from flask import Flask, render_template, request, redirect, flash, session, make_response, jsonify
from config import connect_db
from services import generate_caption
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import os
import re

app = Flask(__name__)
app.secret_key = 'hwstreet_diecast_secret_key_2024'

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
EXCEL_EXTENSIONS = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_excel(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in EXCEL_EXTENSIONS


def to_int(value):
    """Ubah angka dari Excel/string seperti Rp 150.000 menjadi integer."""
    if value is None or str(value).strip() == '':
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()
    text = re.sub(r'[^0-9,.-]', '', text)

    if ',' in text and '.' in text:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '')
    elif ',' in text:
        text = text.replace(',', '')
    elif '.' in text:
        parts = text.split('.')
        if len(parts[-1]) == 3:
            text = text.replace('.', '')

    if text in ('', '-', '.', ','):
        raise ValueError('angka tidak valid')

    return int(float(text))


def rupiah(nilai):
    return "Rp {:,}".format(int(nilai or 0)).replace(',', '.')


def login_required():
    if not session.get('login'):
        return redirect('/login')
    return None


def owner_required():
    auth = login_required()
    if auth:
        return auth

    if session.get('role') != 'owner':
        flash('Halaman ini khusus owner.')
        return redirect('/')

    return None


def catat_history_db(cur, barang_id, nama_barang, aksi,
                     harga_lama=None, harga_baru=None,
                     stok_lama=None, stok_baru=None,
                     keterangan=''):
    """Catat semua aktivitas barang ke tabel history_barang.
    Pakai cursor yang sama supaya ikut commit bareng proses utama.
    """
    cur.execute("""
        INSERT INTO history_barang
        (barang_id, nama_barang, aksi, harga_lama, harga_baru, stok_lama, stok_baru, keterangan, dilakukan_oleh)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        barang_id,
        nama_barang,
        aksi,
        harga_lama,
        harga_baru,
        stok_lama,
        stok_baru,
        keterangan,
        session.get('username', 'admin')
    ))


# =========================
# AUTH
# =========================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('login'):
        return redirect('/')

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = connect_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, username, role
            FROM admin
            WHERE username=%s AND password=%s
        """, (username, password))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user:
            session['login'] = True
            session['username'] = user[1]
            session['role'] = user[2]

            if session['role'] == 'owner':
                return redirect('/laporan')
            return redirect('/')

        flash('Username atau password salah!')
        return render_template('login.html')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# =========================
# DASHBOARD
# =========================

@app.route('/')
def index():
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    cur.execute('SELECT id, nama, merk, kondisi, harga, stok FROM barang')
    barang_rows = cur.fetchall()
    barang = [
        {'id': r[0], 'nama': r[1], 'merk': r[2], 'kondisi': r[3], 'harga': r[4], 'stok': r[5]}
        for r in barang_rows
    ]

    cur.execute("""
        SELECT j.id, b.nama AS barang_nama, j.tanggal
        FROM jadwal j
        JOIN barang b ON j.barang_id = b.id
        ORDER BY j.tanggal DESC
    """)
    jadwal_rows = cur.fetchall()
    jadwal = [{'id': r[0], 'barang_nama': r[1], 'tanggal': r[2]} for r in jadwal_rows]

    cur.close()
    conn.close()

    return render_template('index.html', barang=barang, jadwal=jadwal)


# =========================
# BARANG
# =========================

@app.route('/barang')
def barang():
    auth = login_required()
    if auth:
        return auth

    sort = request.args.get('sort', 'nama_asc')
    order_map = {
        'nama_asc': 'nama ASC, stok DESC',
        'nama_desc': 'nama DESC, stok DESC',
        'stok_desc': 'stok DESC, nama ASC',
        'stok_asc': 'stok ASC, nama ASC',
        'harga_desc': 'harga DESC, nama ASC',
        'harga_asc': 'harga ASC, nama ASC',
        'terbaru': 'id DESC'
    }
    order_by = order_map.get(sort, 'nama ASC, stok DESC')

    conn = connect_db()
    cur = conn.cursor()
    cur.execute(f'SELECT id, nama, merk, kondisi, harga, stok FROM barang ORDER BY {order_by}')
    rows = cur.fetchall()
    cur.close()
    conn.close()

    data = [
        {'id': r[0], 'nama': r[1], 'merk': r[2], 'kondisi': r[3], 'harga': r[4], 'stok': r[5]}
        for r in rows
    ]

    return render_template('barang.html', barang=data, sort=sort)


@app.route('/tambah-barang', methods=['POST'])
def tambah_barang():
    auth = login_required()
    if auth:
        return auth

    nama = request.form['nama']
    merk = request.form['merk']
    kondisi = request.form['kondisi']
    harga = int(request.form['harga'])
    stok = int(request.form['stok'])

    conn = connect_db()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO barang (nama, merk, kondisi, harga, stok)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id
    """, (nama, merk, kondisi, harga, stok))
    barang_id = cur.fetchone()[0]

    catat_history_db(
        cur,
        barang_id,
        nama,
        'Tambah Barang',
        0,
        harga,
        0,
        stok,
        f'Admin menambahkan barang {nama}. Harga awal {rupiah(harga)}, stok awal {stok}.'
    )

    conn.commit()
    cur.close()
    conn.close()

    flash('Barang berhasil ditambahkan dan history tercatat.')
    return redirect('/barang')


@app.route('/import-barang', methods=['POST'])
def import_barang():
    auth = login_required()
    if auth:
        return auth

    if 'file_excel' not in request.files or request.files['file_excel'].filename == '':
        flash('Pilih file Excel terlebih dahulu.')
        return redirect('/barang')

    file = request.files['file_excel']

    if not allowed_excel(file.filename):
        flash('Format file harus .xlsx')
        return redirect('/barang')

    try:
        workbook = load_workbook(file.stream, data_only=True)
        sheet = workbook.active

        data_barang = []
        baris_error = []

        for nomor_baris, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row = list(row or [])
            nama, merk, kondisi, harga, stok = (row + [None] * 5)[:5]

            if all(v is None or str(v).strip() == '' for v in [nama, merk, kondisi, harga, stok]):
                continue

            if not nama or not merk or not kondisi:
                baris_error.append(f'baris {nomor_baris}')
                continue

            try:
                harga_int = to_int(harga)
                stok_int = to_int(stok)
            except ValueError:
                baris_error.append(f'baris {nomor_baris}')
                continue

            data_barang.append((
                str(nama).strip(),
                str(merk).strip(),
                str(kondisi).strip(),
                harga_int,
                stok_int
            ))

        if not data_barang:
            flash('Tidak ada data valid. Pastikan Excel berisi kolom: nama, merk, kondisi, harga, stok.')
            return redirect('/barang')

        conn = connect_db()
        cur = conn.cursor()

        for nama, merk, kondisi, harga, stok in data_barang:
            cur.execute("""
                INSERT INTO barang (nama, merk, kondisi, harga, stok)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (nama, merk, kondisi, harga, stok))
            barang_id = cur.fetchone()[0]

            catat_history_db(
                cur,
                barang_id,
                nama,
                'Import Excel',
                0,
                harga,
                0,
                stok,
                f'Admin mengimport barang {nama} dari Excel. Harga {rupiah(harga)}, stok {stok}.'
            )

        conn.commit()
        cur.close()
        conn.close()

        if baris_error:
            flash(f'{len(data_barang)} barang berhasil diimport. {len(baris_error)} baris dilewati: {", ".join(baris_error[:5])}')
        else:
            flash(f'{len(data_barang)} barang berhasil diimport dari Excel dan history tercatat.')

    except Exception as e:
        flash(f'Gagal import Excel: {e}')

    return redirect('/barang')


@app.route('/update-barang/<int:id>', methods=['POST'])
def update_barang(id):
    auth = login_required()
    if auth:
        return auth

    if request.is_json:
        data = request.get_json()
        nama = data.get('nama')
        merk = data.get('merk')
        kondisi = data.get('kondisi')
        harga = int(data.get('harga') or 0)
        stok = int(data.get('stok') or 0)
    else:
        nama = request.form.get('nama')
        merk = request.form.get('merk')
        kondisi = request.form.get('kondisi')
        harga = to_int(request.form.get('harga') or 0)
        stok = to_int(request.form.get('stok') or 0)

    conn = connect_db()
    cur = conn.cursor()

    cur.execute('SELECT nama, merk, kondisi, harga, stok FROM barang WHERE id=%s', (id,))
    lama = cur.fetchone()

    if not lama:
        cur.close()
        conn.close()
        if request.is_json:
            return jsonify({'status': 'error', 'message': 'Barang tidak ditemukan'}), 404
        flash('Barang tidak ditemukan.')
        return redirect('/barang')

    nama_lama = lama[0]
    harga_lama = int(lama[3] or 0)
    stok_lama = int(lama[4] or 0)

    cur.execute("""
        UPDATE barang
        SET nama=%s, merk=%s, kondisi=%s, harga=%s, stok=%s
        WHERE id=%s
    """, (nama, merk, kondisi, harga, stok, id))

    catat_history_db(
        cur,
        id,
        nama,
        'Update Barang',
        harga_lama,
        harga,
        stok_lama,
        stok,
        f'Admin mengubah {nama_lama}. Harga {rupiah(harga_lama)} menjadi {rupiah(harga)}, stok {stok_lama} menjadi {stok}.'
    )

    conn.commit()
    cur.close()
    conn.close()

    if request.is_json:
        return jsonify({'status': 'ok', 'id': id, 'nama': nama, 'merk': merk, 'kondisi': kondisi, 'harga': harga, 'stok': stok})

    flash('Barang berhasil diperbarui dan history tercatat.')
    return redirect('/barang')


@app.route('/hapus-barang/<int:id>')
def hapus_barang(id):
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    cur.execute('SELECT nama, harga, stok FROM barang WHERE id=%s', (id,))
    barang = cur.fetchone()

    if not barang:
        cur.close()
        conn.close()
        flash('Barang tidak ditemukan.')
        return redirect('/barang')

    nama_barang = barang[0]
    harga_lama = int(barang[1] or 0)
    stok_lama = int(barang[2] or 0)

    catat_history_db(
        cur,
        id,
        nama_barang,
        'Hapus Barang',
        harga_lama,
        0,
        stok_lama,
        0,
        f'Admin menghapus barang {nama_barang}. Harga terakhir {rupiah(harga_lama)}, stok terakhir {stok_lama}.'
    )

    cur.execute('DELETE FROM jadwal WHERE barang_id=%s', (id,))
    cur.execute('DELETE FROM barang WHERE id=%s', (id,))

    conn.commit()
    cur.close()
    conn.close()

    flash('Barang berhasil dihapus dan history tercatat.')
    return redirect('/barang')


@app.route('/hapus-barang-terpilih', methods=['POST'])
def hapus_barang_terpilih():
    auth = login_required()
    if auth:
        return auth

    ids = request.form.getlist('barang_ids')
    ids = [int(i) for i in ids if str(i).isdigit()]

    if not ids:
        flash('Pilih barang yang ingin dihapus terlebih dahulu.')
        return redirect('/barang')

    placeholders = ','.join(['%s'] * len(ids))

    conn = connect_db()
    cur = conn.cursor()

    cur.execute(f'SELECT id, nama, harga, stok FROM barang WHERE id IN ({placeholders})', tuple(ids))
    rows = cur.fetchall()

    for r in rows:
        barang_id = r[0]
        nama_barang = r[1]
        harga_lama = int(r[2] or 0)
        stok_lama = int(r[3] or 0)

        catat_history_db(
            cur,
            barang_id,
            nama_barang,
            'Hapus Barang Terpilih',
            harga_lama,
            0,
            stok_lama,
            0,
            f'Admin menghapus barang terpilih {nama_barang}. Harga terakhir {rupiah(harga_lama)}, stok terakhir {stok_lama}.'
        )

    cur.execute(f'DELETE FROM jadwal WHERE barang_id IN ({placeholders})', tuple(ids))
    cur.execute(f'DELETE FROM barang WHERE id IN ({placeholders})', tuple(ids))

    conn.commit()
    cur.close()
    conn.close()

    flash(f'{len(rows)} barang terpilih berhasil dihapus dan history tercatat.')
    return redirect('/barang')


@app.route('/hapus-semua-barang', methods=['POST'])
def hapus_semua_barang():
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    cur.execute('SELECT id, nama, harga, stok FROM barang')
    rows = cur.fetchall()

    for r in rows:
        barang_id = r[0]
        nama_barang = r[1]
        harga_lama = int(r[2] or 0)
        stok_lama = int(r[3] or 0)

        catat_history_db(
            cur,
            barang_id,
            nama_barang,
            'Hapus Semua Barang',
            harga_lama,
            0,
            stok_lama,
            0,
            f'Admin menghapus semua barang. Item terhapus: {nama_barang}.'
        )

    cur.execute('DELETE FROM jadwal')
    cur.execute('DELETE FROM barang')

    conn.commit()
    cur.close()
    conn.close()

    flash(f'Semua barang berhasil dihapus. Total {len(rows)} history tercatat.')
    return redirect('/barang')


# =========================
# HISTORY OWNER
# =========================

@app.route('/history')
def history():
    auth = owner_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, barang_id, nama_barang, aksi,
               harga_lama, harga_baru, stok_lama, stok_baru,
               keterangan, dilakukan_oleh, waktu
        FROM history_barang
        ORDER BY waktu DESC
    """)
    rows = cur.fetchall()

    data = [
        {
            'id': r[0],
            'barang_id': r[1],
            'nama_barang': r[2],
            'aksi': r[3],
            'harga_lama': r[4],
            'harga_baru': r[5],
            'stok_lama': r[6],
            'stok_baru': r[7],
            'keterangan': r[8],
            'dilakukan_oleh': r[9],
            'waktu': r[10]
        }
        for r in rows
    ]

    cur.close()
    conn.close()

    return render_template('history.html', history=data)


# =========================
# JADWAL LELANG
# =========================

@app.route('/buat-jadwal', methods=['GET'])
def form_jadwal():
    auth = login_required()
    if auth:
        return auth
    return redirect('/jadwal')


@app.route('/buat-jadwal', methods=['POST'])
def buat_jadwal():
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    barang_id = int(request.form['barang_id'])
    tanggal = datetime.fromisoformat(request.form['tanggal'])

    cur.execute('SELECT id, nama, merk, kondisi, harga, stok FROM barang WHERE id=%s', (barang_id,))
    barang = cur.fetchone()

    if not barang:
        cur.close()
        conn.close()
        flash('Barang tidak ditemukan.')
        return redirect('/jadwal')

    nama_barang = barang[1]
    harga_barang = int(barang[4] or 0)
    stok_lama = int(barang[5] or 0)

    if stok_lama <= 0:
        cur.close()
        conn.close()
        flash('Stok habis! Tidak bisa buat jadwal.')
        return redirect('/jadwal')

    if 'poster' not in request.files or request.files['poster'].filename == '':
        cur.close()
        conn.close()
        flash('Tidak ada file poster yang dipilih.')
        return redirect('/jadwal')

    file = request.files['poster']
    if file and allowed_file(file.filename):
        filename = secure_filename(f'poster_{barang_id}_{int(datetime.now().timestamp())}_{file.filename}')
        filepath = os.path.join('static', 'images', filename)
        file.save(filepath)
        image_path = filepath
    else:
        cur.close()
        conn.close()
        flash('Format file tidak didukung. Gunakan PNG, JPG, JPEG, atau GIF.')
        return redirect('/jadwal')

    caption = generate_caption(nama_barang, barang[2], barang[3], harga_barang, tanggal)

    cur.execute("""
        INSERT INTO jadwal (barang_id, tanggal, caption, image)
        VALUES (%s, %s, %s, %s)
    """, (barang_id, tanggal, caption, image_path))

    stok_baru = stok_lama - 1
    cur.execute('UPDATE barang SET stok=%s WHERE id=%s', (stok_baru, barang_id))

    catat_history_db(
        cur,
        barang_id,
        nama_barang,
        'Update Stok Jadwal',
        harga_barang,
        harga_barang,
        stok_lama,
        stok_baru,
        f'Stok {nama_barang} berkurang karena dibuat jadwal lelang. Stok {stok_lama} menjadi {stok_baru}.'
    )

    conn.commit()
    cur.close()
    conn.close()

    flash('Jadwal lelang berhasil dibuat dan stok tercatat di history!')
    return redirect('/jadwal')


@app.route('/jadwal')
def jadwal():
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()

    cur.execute('SELECT id, nama, merk, kondisi, harga, stok FROM barang ORDER BY nama ASC')
    barang_rows = cur.fetchall()

    cur.execute("""
        SELECT j.id, b.nama AS nama, b.merk AS merk, j.tanggal, j.caption, j.image
        FROM jadwal j
        JOIN barang b ON j.barang_id = b.id
        ORDER BY j.id DESC
    """)
    jadwal_rows = cur.fetchall()
    cur.close()
    conn.close()

    barang = [
        {'id': r[0], 'nama': r[1], 'merk': r[2], 'kondisi': r[3], 'harga': r[4], 'stok': r[5]}
        for r in barang_rows
    ]
    data = [
        {'id': r[0], 'nama': r[1], 'merk': r[2], 'tanggal': r[3], 'caption': r[4], 'image': r[5]}
        for r in jadwal_rows
    ]
    return render_template('jadwal.html', jadwal=data, barang=barang)


@app.route('/hapus-jadwal/<int:id>')
def hapus_jadwal(id):
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM jadwal WHERE id=%s', (id,))
    conn.commit()
    cur.close()
    conn.close()

    return redirect('/jadwal')


# =========================
# LAPORAN OWNER
# =========================

@app.route('/laporan')
def laporan():
    auth = owner_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nama, merk, kondisi, harga, stok, (harga * stok) AS total
        FROM barang
        ORDER BY nama ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    data = []
    total_stok = 0
    total_nilai = 0

    for r in rows:
        harga = int(r[4] or 0)
        stok = int(r[5] or 0)
        total = harga * stok
        data.append({'id': r[0], 'nama': r[1], 'merk': r[2], 'kondisi': r[3], 'harga': harga, 'stok': stok, 'total': total})
        total_stok += stok
        total_nilai += total

    return render_template('laporan.html', laporan=data, total_item=len(data), total_stok=total_stok, total_nilai=total_nilai)


@app.route('/laporan/pdf')
def laporan_pdf():
    auth = owner_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT nama, merk, kondisi, harga, stok, (harga * stok) AS total
        FROM barang
        ORDER BY nama ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

    styles = getSampleStyleSheet()
    small_style = ParagraphStyle('SmallText', parent=styles['BodyText'], fontSize=8, leading=10)
    elements = []

    elements.append(Paragraph('Laporan Stok Barang HWStreet Diecast', styles['Title']))
    elements.append(Paragraph(f'Tanggal Cetak: {datetime.now().strftime("%d-%m-%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 12))

    total_stok = sum(int(r[4] or 0) for r in rows)
    total_nilai = sum(int(r[5] or 0) for r in rows)

    elements.append(Paragraph(f'Total Item: {len(rows)}', styles['Normal']))
    elements.append(Paragraph(f'Total Stok: {total_stok}', styles['Normal']))
    elements.append(Paragraph(f'Total Nilai Stok: {rupiah(total_nilai)}', styles['Normal']))
    elements.append(Spacer(1, 12))

    table_data = [['No', 'Nama Barang', 'Merk', 'Kondisi', 'Harga', 'Stok', 'Total']]
    for i, r in enumerate(rows, start=1):
        table_data.append([
            str(i),
            Paragraph(str(r[0]), small_style),
            Paragraph(str(r[1]), small_style),
            str(r[2]),
            rupiah(r[3]),
            str(r[4]),
            rupiah(r[5])
        ])

    table = Table(table_data, repeatRows=1, colWidths=[30, 190, 100, 70, 90, 45, 95])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D0F1C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F8FC')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=laporan_stok_barang.pdf'
    return response


# =========================
# PREVIEW
# =========================

@app.route('/preview/<int:id>')
def preview(id):
    auth = login_required()
    if auth:
        return auth

    conn = connect_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT j.caption, j.image, b.nama
        FROM jadwal j
        JOIN barang b ON j.barang_id = b.id
        WHERE j.id=%s
    """, (id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    data = {'caption': row[0], 'image': row[1], 'nama': row[2]} if row else None
    return render_template('preview.html', data=data)


if __name__ == '__main__':
    app.run(debug=True)
